"""
Lecteur de documents — PDF, Excel, CSV avec metadonnees.
Replique documents-server.mjs de Gemini.
Usage:
  python scripts/doc_reader.py info <fichier>       # Metadonnees
  python scripts/doc_reader.py read <fichier>       # Contenu texte
  python scripts/doc_reader.py read <fichier> --pages 1-3   # Pages PDF specifiques
  python scripts/doc_reader.py read <fichier> --sheet Feuil1 # Feuille Excel specifique
"""
import sys
import os
import argparse
import json

def ensure(pkg, imp=None):
    try:
        return __import__(imp or pkg.replace("-", "_"))
    except ImportError:
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])
        return __import__(imp or pkg.replace("-", "_"))

# --- PDF ---
def pdf_info(path):
    PyPDF2 = ensure("PyPDF2")
    with open(path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        meta = reader.metadata
        return {
            "pages": len(reader.pages),
            "title": getattr(meta, "title", None),
            "author": getattr(meta, "author", None),
            "creator": getattr(meta, "creator", None),
            "producer": getattr(meta, "producer", None),
        }

def pdf_read(path, pages=None):
    PyPDF2 = ensure("PyPDF2")
    with open(path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        total = len(reader.pages)
        if pages:
            start, end = (int(x) for x in pages.split("-")) if "-" in pages else (int(pages), int(pages))
            start, end = max(1, start) - 1, min(total, end)
            indices = range(start, end)
        else:
            indices = range(total)
        text = ""
        for i in indices:
            page_text = reader.pages[i].extract_text() or ""
            text += f"\n--- Page {i+1} ---\n{page_text}"
    return text.strip()

# --- Excel ---
def excel_info(path):
    ensure("openpyxl")
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({"name": name, "rows": ws.max_row, "cols": ws.max_column})
    wb.close()
    return {"sheets": sheets}

def excel_read(path, sheet=None, max_rows=100):
    pd = ensure("pandas")
    ensure("openpyxl")
    kwargs = {"engine": "openpyxl", "nrows": max_rows}
    if sheet:
        kwargs["sheet_name"] = sheet
    df = pd.read_excel(path, **kwargs)
    return df.to_string(index=False, max_rows=max_rows)

# --- CSV ---
def csv_info(path):
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        first_line = f.readline()
        delimiter = ";" if ";" in first_line else ","
        f.seek(0)
        lines = sum(1 for _ in f)
        cols = len(first_line.split(delimiter))
    return {"rows": lines - 1, "cols": cols, "delimiter": delimiter}

def csv_read(path, max_rows=100):
    pd = ensure("pandas")
    info = csv_info(path)
    df = pd.read_csv(path, sep=info["delimiter"], nrows=max_rows)
    return df.to_string(index=False, max_rows=max_rows)

# --- Dispatch ---
READERS = {
    ".pdf": {"info": pdf_info, "read": pdf_read},
    ".xlsx": {"info": excel_info, "read": excel_read},
    ".xls": {"info": excel_info, "read": excel_read},
    ".csv": {"info": csv_info, "read": csv_read},
    ".tsv": {"info": csv_info, "read": csv_read},
}

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("action", choices=["info", "read"])
    parser.add_argument("file")
    parser.add_argument("--pages", default=None, help="Plage de pages PDF (ex: 1-3)")
    parser.add_argument("--sheet", default=None, help="Nom de la feuille Excel")
    parser.add_argument("--max-rows", type=int, default=100)
    args = parser.parse_args()

    ext = os.path.splitext(args.file)[1].lower()
    if ext not in READERS:
        print(f"Format non supporte: {ext}. Supportes: {', '.join(READERS.keys())}")
        sys.exit(1)

    handler = READERS[ext][args.action]
    if args.action == "read" and ext == ".pdf":
        result = handler(args.file, pages=args.pages)
    elif args.action == "read" and ext in (".xlsx", ".xls"):
        result = handler(args.file, sheet=args.sheet, max_rows=args.max_rows)
    else:
        result = handler(args.file)

    if isinstance(result, (dict, list)):
        print(json.dumps(result, indent=2, ensure_ascii=False))
    else:
        print(result)
